"""Exports records and summary to an Excel file — styled with openpyxl."""
import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.record import Record, VoteStatus
from app.models.user import User

logger = logging.getLogger(__name__)

# ── Colour tokens ─────────────────────────────────────────────────────────────
_C = {
    "hdr_bg":      "1F3864",   # header row — dark navy
    "hdr_fg":      "FFFFFF",
    "name_bg_odd": "FFFFFF",   # participant name rows
    "name_bg_eve": "F2F4F9",
    "sep_line":    "4472C4",   # thick separator above stats
    "stat_bg":     "DAE8FC",   # micro-stats rows
    "stat_fg":     "1A2E5A",
    "sum_hdr_bg":  "1F3864",   # summary header (same as matrix)
    "sum_alt":     "EEF2FA",   # summary alternating row
}

_STATUS_FMT: dict[VoteStatus, tuple[str, str, bool]] = {
    # (bg, fg, bold)
    VoteStatus.YES:     ("C6EFCE", "276221", True),
    VoteStatus.NO:      ("FFC7CE", "9C0006", True),
    VoteStatus.ORG:     ("E2D9F3", "5B2C8D", True),
    VoteStatus.UNKNOWN: ("FFEB9C", "7D5A00", False),
    VoteStatus.NA:      ("DDDDDD", "555555", False),
}

# ── Style helpers ─────────────────────────────────────────────────────────────
_FONT = "Calibri"
_FONT_SZ = 11


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _side(style: str = "thin", color: str = "BFBFBF") -> Side:
    return Side(style=style, color=color)


def _border_thin() -> Border:
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _border_sep_top() -> Border:
    """First stats row: thick top border to visually separate from participants."""
    thick = _side("medium", _C["sep_line"])
    thin = _side()
    return Border(left=thin, right=thin, top=thick, bottom=thin)


def _align(h: str = "center", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _hdr_cell(cell, text: str) -> None:
    cell.value = text
    cell.fill = _fill(_C["hdr_bg"])
    cell.font = Font(bold=True, color=_C["hdr_fg"], size=_FONT_SZ, name=_FONT)
    cell.alignment = _align("center", "center", wrap=True)
    cell.border = _border_thin()


def _name_cell(cell, text: str, row_even: bool) -> None:
    cell.value = text
    cell.fill = _fill(_C["name_bg_eve"] if row_even else _C["name_bg_odd"])
    cell.font = Font(size=_FONT_SZ, name=_FONT)
    cell.alignment = _align("left")
    cell.border = _border_thin()


def _status_cell(cell, status: VoteStatus) -> None:
    bg, fg, bold = _STATUS_FMT[status]
    cell.value = status.value
    cell.fill = _fill(bg)
    cell.font = Font(bold=bold, color=fg, size=_FONT_SZ, name=_FONT)
    cell.alignment = _align("center")
    cell.border = _border_thin()


def _stat_cell(cell, value, is_label: bool, first_stat_row: bool) -> None:
    cell.value = value
    cell.fill = _fill(_C["stat_bg"])
    cell.font = Font(
        bold=is_label, color=_C["stat_fg"], size=_FONT_SZ, name=_FONT
    )
    cell.alignment = _align("left" if is_label else "center")
    cell.border = _border_sep_top() if first_stat_row else _border_thin()


def _set_col_width(ws, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


# ── Sheet builders ────────────────────────────────────────────────────────────
def _build_matrix(
    ws,
    records: list[Record],
    members: dict[int, User],
) -> None:
    dates = sorted({r.poll_date.date() for r in records})
    date_labels = [d.strftime("%d.%m") for d in dates]
    n_cols = 1 + len(dates)

    # ── Row 1: headers ──
    _hdr_cell(ws.cell(1, 1), "Участник")
    for ci, lbl in enumerate(date_labels, 2):
        _hdr_cell(ws.cell(1, ci), lbl)
    ws.row_dimensions[1].height = 22

    # ── Rows 2…N+1: participants ──
    status_map: dict[tuple[int, object], VoteStatus] = {
        (r.user_id, r.poll_date.date()): r.status for r in records
    }
    for ri, (uid, user) in enumerate(members.items(), 2):
        row_even = (ri % 2 == 0)
        _name_cell(ws.cell(ri, 1), user.full_name, row_even)
        for ci, d in enumerate(dates, 2):
            status = status_map.get((uid, d), VoteStatus.UNKNOWN)
            _status_cell(ws.cell(ri, ci), status)
        ws.row_dimensions[ri].height = 20

    # ── Micro-stats rows ──
    stat_defs = [
        (VoteStatus.YES,     "Придут"),
        (VoteStatus.NO,      "Не придут"),
        (VoteStatus.UNKNOWN, "Не ответили"),
        (VoteStatus.ORG,     "Организаторы"),
        (VoteStatus.NA,      "Нет данных"),
    ]
    n_members = len(members)
    for si, (vote_status, label) in enumerate(stat_defs):
        ri = 2 + n_members + si
        first = si == 0
        _stat_cell(ws.cell(ri, 1), label, is_label=True, first_stat_row=first)
        for ci, d in enumerate(dates, 2):
            count = sum(
                1 for r in records
                if r.poll_date.date() == d and r.status == vote_status
            )
            _stat_cell(ws.cell(ri, ci), count, is_label=False, first_stat_row=first)
        ws.row_dimensions[ri].height = 20

    # ── Column widths ──
    _set_col_width(ws, 1, 28)
    for ci in range(2, n_cols + 1):
        _set_col_width(ws, ci, 7)

    # ── Freeze header + name column ──
    ws.freeze_panes = ws.cell(2, 2)


def _build_summary(ws, summary: list[dict]) -> None:
    col_map = [
        ("user",      "Участник"),
        ("attended",  "Был (+)"),
        ("missed",    "Не был (−)"),
        ("org",       "Организатор"),
        ("unknown",   "Не ответил (/)"),
        ("na",        "Нет данных"),
        ("total",     "Всего"),
    ]
    col_keys  = [k for k, _ in col_map]
    col_names = [n for _, n in col_map]

    # ── Row 1: headers ──
    for ci, name in enumerate(col_names, 1):
        _hdr_cell(ws.cell(1, ci), name)
    ws.row_dimensions[1].height = 22

    # ── Data rows ──
    for ri, row in enumerate(summary, 2):
        row_even = (ri % 2 == 0)
        bg = _C["sum_alt"] if row_even else "FFFFFF"
        for ci, key in enumerate(col_keys, 1):
            cell = ws.cell(ri, ci, value=row.get(key, ""))
            cell.fill = _fill(bg)
            cell.font = Font(size=_FONT_SZ, name=_FONT)
            cell.alignment = _align("left" if ci == 1 else "center")
            cell.border = _border_thin()
        ws.row_dimensions[ri].height = 20

    # ── Column widths ──
    widths = [28, 9, 12, 13, 16, 12, 9]
    for ci, w in enumerate(widths, 1):
        _set_col_width(ws, ci, w)

    ws.freeze_panes = ws.cell(2, 1)


# ── Public entry point ────────────────────────────────────────────────────────
def export(
    records: list[Record],
    members: dict[int, User],
    summary: list[dict],
    output_path: Path,
) -> None:
    wb = Workbook()

    ws_matrix = wb.active
    ws_matrix.title = "Матрица"
    _build_matrix(ws_matrix, records, members)

    ws_summary = wb.create_sheet("Сводка")
    _build_summary(ws_summary, summary)

    wb.save(output_path)
    logger.info("Report saved to %s", output_path)
